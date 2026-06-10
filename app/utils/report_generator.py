from io import BytesIO
from datetime import date, datetime
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


_HEADER_BG = "1E3A5F"
_ALT_ROW_BG = "F1F5F9"
_BORDER_COLOR = "CBD5E1"

_THIN = Side(style="thin", color=_BORDER_COLOR)
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

COLUMNS = [
    ("Fecha creación", 20),
    ("No. Requisición", 22),
    ("Proveedor", 35),
    ("Subtotal", 18),
    ("Cuentas", 40),
    ("Proyecto", 35),
    ("Estatus", 18),
]

STATUS_LABELS = {
    "pending":   "Pendiente",
    "approved":  "Aprobada",
    "rejected":  "Rechazada",
    "cancelled": "Cancelada",
    "draft":     "Borrador",
}


def _header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _CELL_BORDER


def _data_style(cell, alt_row: bool):
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = _CELL_BORDER
    if alt_row:
        cell.fill = PatternFill("solid", fgColor=_ALT_ROW_BG)


def _subtotal_mxn(req) -> float:
    subtotal = float(req.subtotal or 0)
    rate = float(req.exchange_rate or 1)
    currency = req.currency or "MXN"
    if currency == "USD" and rate <= 1:
        return subtotal  # no meaningful rate — keep USD value
    return subtotal * rate


def generate_requisitions_report(
    requisitions,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Requisiciones"

    ncols = len(COLUMNS)

    # ── row 1: company title ──────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title_cell = ws.cell(row=1, column=1, value="INGENIERIA EN MANUFACTURAS Y SERVICIOS SA DE CV")
    title_cell.font = Font(bold=True, size=14, color="1E3A5F")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── row 2: subtitle ───────────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    subtitle_cell = ws.cell(row=2, column=1, value="Movimientos Auxiliares por Segmento de Negocio")
    subtitle_cell.font = Font(bold=True, size=12, color="2563EB")
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # ── row 3: date range from filters ───────────────────────────────────────
    from_str = date_from.strftime("%d/%m/%Y") if date_from else "—"
    to_str = date_to.strftime("%d/%m/%Y") if date_to else "—"
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    date_range_cell = ws.cell(row=3, column=1, value=f"Período: {from_str}  —  {to_str}")
    date_range_cell.font = Font(size=10, color="475569")
    date_range_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18

    # ── row 4: generated timestamp ────────────────────────────────────────────
    generated_cell = ws.cell(
        row=4, column=1,
        value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    generated_cell.font = Font(italic=True, size=9, color="64748B")
    ws.row_dimensions[4].height = 16

    # ── headers ──────────────────────────────────────────────────────────────
    HEADER_ROW = 6
    for col_idx, (label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
        _header_style(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[HEADER_ROW].height = 22

    # ── data rows ─────────────────────────────────────────────────────────────
    for row_idx, req in enumerate(requisitions, start=HEADER_ROW + 1):
        alt = (row_idx % 2 == 0)

        accounts = _collect_accounts(req)

        created_at = req.created_at
        if created_at and hasattr(created_at, "strftime"):
            created_at_str = created_at.strftime("%d/%m/%Y")
        else:
            created_at_str = str(created_at) if created_at else ""

        row_data = [
            created_at_str,
            req.requisition_number,
            req.supplier.name if req.supplier else "",
            _subtotal_mxn(req),
            accounts,
            req.project.name if req.project else "",
            STATUS_LABELS.get(req.status, req.status),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _data_style(cell, alt)

            # currency format for subtotal column (index 4)
            if col_idx == 4:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

        ws.row_dimensions[row_idx].height = 18

    # ── freeze header ─────────────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


EXPORT_COLUMNS = [
    ("ID",              8),
    ("No. Requisición", 22),
    ("Proveedor",       35),
    ("Estatus",         18),
    ("Moneda",          10),
    ("Total",           18),
    ("Creado por",      28),
    ("Fecha creación",  20),
]


def generate_requisitions_export(requisitions) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Requisiciones"

    HEADER_ROW = 1
    for col_idx, (label, width) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
        _header_style(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[HEADER_ROW].height = 22

    for row_idx, req in enumerate(requisitions, start=HEADER_ROW + 1):
        alt = (row_idx % 2 == 0)

        created_at = req.created_at
        created_at_str = created_at.strftime("%d/%m/%Y") if created_at and hasattr(created_at, "strftime") else str(created_at or "")

        row_data = [
            req.id,
            req.requisition_number,
            req.supplier.supplier_code if req.supplier else "",
            STATUS_LABELS.get(req.status, req.status),
            req.currency or "MXN",
            float(req.total_amount or 0),
            req.creator.name if req.creator else "",
            created_at_str,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _data_style(cell, alt)
            if col_idx == 6:  # Total
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


LIST_EXCEL_COLUMNS = [
    ("No. Requisición",  22),
    ("Proyecto",         35),
    ("Proveedor",        35),
    ("Estatus",          16),
    ("Moneda",           10),
    ("Total",            18),
    ("OC",               20),
    ("Creado por",       28),
    ("Fecha creación",   20),
]


def generate_requisitions_list_excel(requisitions) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Requisiciones"

    # header row
    HEADER_ROW = 1
    for col_idx, (label, width) in enumerate(LIST_EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
        _header_style(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[HEADER_ROW].height = 22

    for row_idx, req in enumerate(requisitions, start=HEADER_ROW + 1):
        alt = (row_idx % 2 == 0)

        created_at = req.created_at
        created_at_str = created_at.strftime("%d/%m/%Y") if created_at and hasattr(created_at, "strftime") else str(created_at or "")

        row_data = [
            req.requisition_number,
            req.project.name if req.project else "",
            req.supplier.name if req.supplier else "",
            STATUS_LABELS.get(req.status, req.status),
            req.currency or "MXN",
            float(req.total_amount or 0),
            req.purchase_order or "",
            req.creator.name if req.creator else "",
            created_at_str,
        ]

        CURRENCY_COLS = {6}  # 1-based index of Total column

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _data_style(cell, alt)
            if col_idx in CURRENCY_COLS:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _collect_accounts(req) -> str:
    seen = set()
    parts = []
    for item in getattr(req, "items", []):
        account = getattr(item, "account", None)
        if account and account.id not in seen:
            seen.add(account.id)
            parts.append(account.account_number)
    return ", ".join(parts)
